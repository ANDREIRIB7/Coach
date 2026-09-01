"""
Coach de Estudos — versão 100% Python (Streamlit + Google Drive)

Roda como site (Streamlit Community Cloud, Render, etc.) e guarda os dados
em um único arquivo JSON numa pasta do seu Google Drive — sem depender de
nenhum banco que possa "hibernar".

App pessoal, sem tela de login: qualquer pessoa com o link do app acessa
os mesmos dados.
"""

import io
import json
import math
import re
import unicodedata
import uuid
from datetime import date, datetime, timedelta

import pandas as pd
import streamlit as st
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaIoBaseUpload

# ----------------------------------------------------------------------------
# Configuração da página e do cliente do Google Drive
# ----------------------------------------------------------------------------

st.set_page_config(page_title="Coach de Estudos", page_icon="🎯", layout="wide")

APP_OWNER_NAME = "Andrei"
DRIVE_SCOPES = ["https://www.googleapis.com/auth/drive"]
GDRIVE_FOLDER_ID = st.secrets.get("GDRIVE_FOLDER_ID", "")
STATE_FILENAME = "coach_estudos_state.json"


@st.cache_resource
def get_drive_service():
    sa_info = st.secrets.get("gdrive_service_account")
    if not sa_info or not GDRIVE_FOLDER_ID:
        return None
    creds = service_account.Credentials.from_service_account_info(dict(sa_info), scopes=DRIVE_SCOPES)
    return build("drive", "v3", credentials=creds, cache_discovery=False)


drive = get_drive_service()

# ----------------------------------------------------------------------------
# Tema visual (paleta inspirada no layout de referência enviado)
# ----------------------------------------------------------------------------

ACCENT = "#2F6FEE"
ACCENT_SOFT = "#EAF1FF"
GREEN = "#22C55E"
PURPLE = "#7C5CFC"
AMBER = "#F59E0B"
BG = "#EEF1F8"
CARD = "#FFFFFF"
BORDER = "#E7E9F3"
TEXT = "#1E2432"
MUTED = "#8A93A6"
PALETTE = [ACCENT, GREEN, PURPLE, AMBER, "#EC4899", "#14B8A6"]

CUSTOM_CSS = f"""
<style>
.stApp {{ background: {BG}; }}
[data-testid="stSidebar"] {{
    background: {CARD};
    border-right: 1px solid {BORDER};
}}
[data-testid="stSidebar"] .block-container {{ padding-top: 1.4rem; }}

/* menus de navegação em formato de pílula (páginas e editais) */
div[data-testid="stRadio"] > div[role="radiogroup"] {{ gap: 3px; }}
div[data-testid="stRadio"] label {{
    border-radius: 10px;
    padding: 9px 12px;
    width: 100%;
    transition: background 0.15s ease;
}}
div[data-testid="stRadio"] label:hover {{ background: {ACCENT_SOFT}; }}
div[data-testid="stRadio"] label:has(input:checked) {{ background: {ACCENT}; }}
div[data-testid="stRadio"] label:has(input:checked) p {{ color: #fff !important; font-weight: 600; }}
div[data-testid="stRadio"] label > div:first-child {{ display: none; }}

/* cartões (qualquer container com borda vira um "card" no estilo do painel) */
[data-testid="stVerticalBlockBorderWrapper"] {{
    background: {CARD};
    border: 1px solid {BORDER} !important;
    border-radius: 18px !important;
    box-shadow: 0 1px 3px rgba(15,23,42,0.06);
}}

.topbar-search input {{
    border-radius: 999px !important;
    background: {BG} !important;
    border: 1px solid {BORDER} !important;
}}
.avatar-chip {{
    width: 40px; height: 40px; border-radius: 50%;
    background: {ACCENT}; color: #fff; display: flex;
    align-items: center; justify-content: center; font-weight: 700; font-size: 15px;
}}
.hero-eyebrow {{ color: {MUTED}; font-size: 12.5px; margin-bottom: 4px; }}
.hero-number {{ font-size: 38px; font-weight: 800; color: {TEXT}; line-height: 1.15; }}
.hero-delta {{ display:inline-block; background:{ACCENT_SOFT}; color:{ACCENT}; font-size:12.5px;
    font-weight:600; padding:2px 9px; border-radius:999px; margin-left:8px; }}
.metric-dot {{ display:inline-block; width:9px; height:9px; border-radius:50%; margin-right:6px; }}
.subject-icon {{
    width: 42px; height: 42px; border-radius: 12px; display: flex;
    align-items: center; justify-content: center; font-weight: 700; color: #fff; font-size: 16px;
}}
.today-card {{ background: {ACCENT_SOFT}; border-radius: 18px; padding: 20px 22px; }}
.today-card .hero-number {{ color: {ACCENT}; }}
.coach-card {{
    border: 1px solid {ACCENT};
    background: {ACCENT_SOFT};
    border-radius: 16px;
    padding: 16px 20px;
    margin-bottom: 18px;
}}
.coach-eyebrow {{ font-size: 11px; letter-spacing: 0.08em; color: {ACCENT}; margin-bottom: 6px; font-weight: 700; }}
.coach-subject {{ font-size: 19px; font-weight: 700; margin-bottom: 2px; color: {TEXT}; }}
.coach-topic {{ color: {MUTED}; font-size: 13px; margin-bottom: 8px; }}
.coach-reason {{ display: inline-block; font-size: 11.5px; background: #fff; border-radius: 12px;
    padding: 3px 10px; color: {ACCENT}; border: 1px solid {ACCENT}; }}
.badge-earned {{ display:inline-block; padding: 4px 10px; margin: 2px; border-radius: 12px;
    border: 1px solid {ACCENT}; color:{ACCENT}; font-size: 12px; }}
.badge-locked {{ display:inline-block; padding: 4px 10px; margin: 2px; border-radius: 12px;
    border: 1px solid #ccc; color:#999; font-size: 12px; }}
.class-badge {{ display:inline-block; font-size: 11px; font-weight: 600; padding: 2px 9px;
    border-radius: 999px; margin-left: 8px; vertical-align: middle; }}
.class-geral {{ background: {ACCENT_SOFT}; color: {ACCENT}; }}
.class-especifico {{ background: #FCE9FF; color: {PURPLE}; }}
</style>
"""
st.markdown(CUSTOM_CSS, unsafe_allow_html=True)


def subject_color(name: str) -> str:
    return PALETTE[hash(name or "") % len(PALETTE)]


# ----------------------------------------------------------------------------
# Funções utilitárias
# ----------------------------------------------------------------------------


def new_id() -> str:
    return uuid.uuid4().hex[:10]


def iso_week_key(d: date) -> str:
    y, w, _ = d.isocalendar()
    return f"{y}-W{w:02d}"


def week_info(offset: int):
    today = date.today()
    monday = today - timedelta(days=today.weekday()) + timedelta(weeks=offset)
    sunday = monday + timedelta(days=6)
    return {
        "key": iso_week_key(monday),
        "monday": monday,
        "sunday": sunday,
        "label": f"{monday.strftime('%d/%m')} – {sunday.strftime('%d/%m')}",
    }


def days_until(date_str):
    if not date_str:
        return None
    try:
        target = datetime.strptime(date_str, "%Y-%m-%d").date()
    except ValueError:
        return None
    return (target - date.today()).days


def subject_key(materia: str, assunto: str) -> str:
    return f"{(materia or '').strip()}||{(assunto or '').strip()}"


def compute_streak(daily_activity: dict) -> int:
    streak = 0
    cursor = date.today()
    if daily_activity.get(cursor.isoformat(), 0) <= 0:
        cursor -= timedelta(days=1)
    while daily_activity.get(cursor.isoformat(), 0) > 0:
        streak += 1
        cursor -= timedelta(days=1)
    return streak


def level_from_xp(xp: int):
    level = 1 + int(math.sqrt(xp / 40)) if xp > 0 else 1
    floor = 40 * (level - 1) ** 2
    ceil = 40 * level ** 2
    progress = (xp - floor) / (ceil - floor) if ceil > floor else 0.0
    return level, progress


def weekly_goal_for(subject: dict, total_goal: int, sum_weights: float) -> int:
    if not total_goal or not sum_weights:
        return 0
    return max(1, round(total_goal * subject.get("peso", 1) / sum_weights))


def _norm(s) -> str:
    """minúsculas, sem acento, sem espaços nas pontas — para comparar cabeçalhos"""
    s = str(s or "").strip().lower()
    return "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))


def parse_tec_counter(text: str):
    """Extrai (respondidas, acertos, erros) de um texto colado do contador do
    TEC Concursos, ex.: "1 de 964 (1 R, 1 A e 0 E)". Retorna None se não
    conseguir reconhecer o padrão."""
    if not text:
        return None
    m = re.search(r"\(([^)]*)\)", text)
    inner = m.group(1) if m else text
    r_m = re.search(r"(\d+)\s*R\b", inner, re.IGNORECASE)
    a_m = re.search(r"(\d+)\s*A\b", inner, re.IGNORECASE)
    e_m = re.search(r"(\d+)\s*E\b", inner, re.IGNORECASE)
    if not r_m or not a_m:
        return None
    respondidas = int(r_m.group(1))
    acertos = int(a_m.group(1))
    erros = int(e_m.group(1)) if e_m else max(respondidas - acertos, 0)
    return respondidas, acertos, erros


# ----------------------------------------------------------------------------
# Importador de planilhas — reconhece o padrão "Edital Verticalizado"
# ----------------------------------------------------------------------------

HEADER_ALIASES = {
    "materia": ["materia", "disciplina", "grupo / disciplina", "grupo/disciplina", "grupo"],
    "assunto": [
        "assunto", "topico", "tema",
        "conteudo programatico / topico", "conteudo programatico/topico", "conteudo programatico",
    ],
    "peso": ["peso", "importancia", "prioridade"],
    "classificacao": ["classificacao", "classificacao geral/especifico", "tipo", "categoria", "geral/especifico"],
    "link": ["linktec", "link tec", "link do tec", "link", "tec"],
    "done": ["questoes feitas", "feitas"],
    "accuracy": ["% acertos", "acertos", "% acerto"],
}


def _normalize_classificacao(raw) -> str:
    v = _norm(raw)
    if not v:
        return "Geral"
    if "espec" in v:
        return "Específico"
    return "Geral"


def parse_verticalizado(df_raw: pd.DataFrame):
    """Recebe uma planilha lida SEM cabeçalho (header=None). Localiza a linha
    de cabeçalho de verdade (pulando título/subtítulo) e ignora as linhas de
    bloco (ex.: "CONHECIMENTOS BÁSICOS - P1 (35 itens...)") que só preenchem
    a coluna de Matéria/Disciplina. Retorna None se não reconhecer o formato."""
    header_row_idx = None
    col_map = {}
    for i in range(min(15, len(df_raw))):
        row = df_raw.iloc[i]
        found = {}
        for j, cell in enumerate(row):
            cell_norm = _norm(cell)
            if not cell_norm:
                continue
            for field, aliases in HEADER_ALIASES.items():
                if field not in found and cell_norm in aliases:
                    found[field] = j
        if "materia" in found and "assunto" in found:
            header_row_idx = i
            col_map = found
            break
    if header_row_idx is None:
        return None

    imported = []
    current_block_class = "Geral"
    for i in range(header_row_idx + 1, len(df_raw)):
        row = df_raw.iloc[i]
        materia_val = row[col_map["materia"]]
        assunto_val = row[col_map["assunto"]]
        materia = str(materia_val).strip() if pd.notna(materia_val) else ""
        assunto = str(assunto_val).strip() if pd.notna(assunto_val) else ""
        if not materia or materia.lower() == "nan":
            continue
        if not assunto or assunto.lower() == "nan":
            # linha de bloco/seção (ex.: "CONHECIMENTOS BÁSICOS - P1..."). Não é uma matéria de
            # verdade, mas o texto costuma indicar se o bloco é geral ou específico do cargo —
            # usamos isso como padrão para as linhas seguintes, até o próximo bloco.
            block_norm = _norm(materia)
            if "especific" in block_norm or "especializ" in block_norm:
                current_block_class = "Específico"
            elif "basic" in block_norm or "geral" in block_norm:
                current_block_class = "Geral"
            continue

        peso = 3.0
        if "peso" in col_map and pd.notna(row[col_map["peso"]]):
            try:
                peso = float(row[col_map["peso"]])
            except (TypeError, ValueError):
                pass

        if "classificacao" in col_map and pd.notna(row[col_map["classificacao"]]):
            classificacao = _normalize_classificacao(row[col_map["classificacao"]])
        else:
            classificacao = current_block_class

        link = ""
        if "link" in col_map and pd.notna(row[col_map["link"]]):
            link = str(row[col_map["link"]]).strip()

        done = 0
        if "done" in col_map and pd.notna(row[col_map["done"]]):
            try:
                done = int(float(row[col_map["done"]]))
            except (TypeError, ValueError):
                pass

        acc = None
        if "accuracy" in col_map and pd.notna(row[col_map["accuracy"]]):
            try:
                acc = float(row[col_map["accuracy"]])
                if acc > 1:
                    acc = acc / 100
            except (TypeError, ValueError):
                pass

        correct = int(round(done * acc)) if (done and acc is not None) else 0
        imported.append(
            {
                "materia": materia, "assunto": assunto, "peso": peso, "classificacao": classificacao,
                "link": link, "done": done, "correct": correct,
            }
        )
    return imported


def build_template_xlsx() -> bytes:
    """Gera o modelo de planilha (.xlsx) para download, já formatado e com
    a coluna de Classificação (Geral/Específico) com lista suspensa."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Edital"

    headers = ["Matéria", "Assunto", "Peso", "Classificação", "Link TEC"]
    widths = [30, 55, 10, 16, 40]

    title_font = Font(name="Arial", size=14, bold=True, color="1E2432")
    note_font = Font(name="Arial", size=10, italic=True, color="8A93A6")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F6FEE")
    example_font = Font(name="Arial", size=10, italic=True, color="8A93A6")
    body_font = Font(name="Arial", size=10, color="1E2432")
    thin = Side(style="thin", color="E7E9F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:E1")
    ws["A1"] = "Modelo de Edital Verticalizado — Coach de Estudos"
    ws["A1"].font = title_font

    ws.merge_cells("A2:E2")
    ws["A2"] = (
        "Preencha uma linha por tópico. Peso vai de 1 (baixa prioridade) a 5 (altíssima prioridade). "
        "Classificação aceita apenas Geral (conhecimentos comuns/básicos) ou Específico (do cargo). "
        "Link TEC é opcional."
    )
    ws["A2"].font = note_font
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[2].height = 30

    header_row = 4
    for j, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=j, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    examples = [
        ("Direito Constitucional", "Controle de Constitucionalidade", 5, "Específico", ""),
        ("Língua Portuguesa", "Concordância Verbal e Nominal", 3, "Geral", ""),
    ]
    for i, row_vals in enumerate(examples, start=header_row + 1):
        for j, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.font = example_font
            cell.border = border
            if j == 2:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    last_row = header_row + 200
    for i in range(header_row + len(examples) + 1, last_row + 1):
        for j in range(1, len(headers) + 1):
            ws.cell(row=i, column=j).font = body_font
            ws.cell(row=i, column=j).border = border

    dv_class = DataValidation(type="list", formula1='"Geral,Específico"', allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv_class)
    dv_class.add(f"D{header_row + 1}:D{last_row}")

    dv_peso = DataValidation(type="whole", operator="between", formula1=1, formula2=5, allow_blank=True)
    dv_peso.error = "Peso deve ser um número inteiro de 1 a 5."
    ws.add_data_validation(dv_peso)
    dv_peso.add(f"C{header_row + 1}:C{last_row}")

    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = w

    ws.freeze_panes = f"A{header_row + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()



# ----------------------------------------------------------------------------
# Persistência (Google Drive — um único arquivo JSON)
# ----------------------------------------------------------------------------

DEFAULT_STATE = {"editais": [], "currentEditalId": None, "weeklyLog": {}, "dailyActivity": {}}


def _find_state_file_id():
    query = f"'{GDRIVE_FOLDER_ID}' in parents and name = '{STATE_FILENAME}' and trashed = false"
    res = drive.files().list(q=query, fields="files(id, name)", spaces="drive").execute()
    files = res.get("files", [])
    return files[0]["id"] if files else None


def load_state() -> dict:
    try:
        file_id = _find_state_file_id()
        if not file_id:
            # ainda não existe — será criado no primeiro save_state()
            return {k: (type(v)() if isinstance(v, (list, dict)) else v) for k, v in DEFAULT_STATE.items()}
        request = drive.files().get_media(fileId=file_id)
        buf = io.BytesIO()
        downloader = MediaIoBaseDownload(buf, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()
        buf.seek(0)
        loaded = json.loads(buf.read().decode("utf-8"))
        for k, v in DEFAULT_STATE.items():
            loaded.setdefault(k, v if not isinstance(v, (list, dict)) else type(v)())
        st.session_state["_gdrive_file_id"] = file_id
        return loaded
    except Exception as e:
        st.warning(f"Não foi possível carregar seus dados salvos: {e}")
        return {k: (type(v)() if isinstance(v, (list, dict)) else v) for k, v in DEFAULT_STATE.items()}


def save_state(state: dict):
    try:
        payload = json.dumps(state, ensure_ascii=False).encode("utf-8")
        media = MediaIoBaseUpload(io.BytesIO(payload), mimetype="application/json", resumable=False)
        file_id = st.session_state.get("_gdrive_file_id") or _find_state_file_id()
        if file_id:
            drive.files().update(fileId=file_id, media_body=media).execute()
        else:
            metadata = {"name": STATE_FILENAME, "parents": [GDRIVE_FOLDER_ID]}
            created = drive.files().create(body=metadata, media_body=media, fields="id").execute()
            file_id = created["id"]
        st.session_state["_gdrive_file_id"] = file_id
    except Exception as e:
        st.warning(f"Não foi possível salvar agora — verifique sua conexão. ({e})")


if drive is None:
    st.error(
        "Faltam as credenciais do Google Drive. Configure `GDRIVE_FOLDER_ID` e "
        "`[gdrive_service_account]` em `.streamlit/secrets.toml` (veja o README.md)."
    )
    st.stop()

if "state" not in st.session_state:
    st.session_state.state = load_state()

state = st.session_state.state
user_label = APP_OWNER_NAME

PAGES = ["Painel", "Matérias & Pesos", "Importar planilha", "Dashboard"]
if "page" not in st.session_state:
    st.session_state.page = "Painel"

# ----------------------------------------------------------------------------
# Barra lateral — navegação e editais
# ----------------------------------------------------------------------------

with st.sidebar:
    st.markdown("### 🎯 Coach de Estudos")
    st.session_state.page = st.radio(
        "Navegação",
        PAGES,
        index=PAGES.index(st.session_state.page),
        label_visibility="collapsed",
        key="page_radio",
    )

    st.divider()
    st.markdown("**📚 Editais**")
    editais = state["editais"]

    current_edital = None
    if editais:
        names = [e["name"] for e in editais]
        idx_default = next((i for i, e in enumerate(editais) if e["id"] == state.get("currentEditalId")), 0)
        selected_name = st.radio("Selecione", names, index=idx_default, label_visibility="collapsed", key="edital_radio")
        current_edital = next(e for e in editais if e["name"] == selected_name)
        state["currentEditalId"] = current_edital["id"]

    with st.expander("➕ Novo edital"):
        new_name = st.text_input("Nome", key="new_edital_name", placeholder="Ex: TCDF 2026")
        new_date = st.date_input("Data da prova (opcional)", value=None, key="new_edital_date")
        new_goal = st.number_input("Meta semanal total de questões", min_value=0, value=100, key="new_edital_goal")
        if st.button("Criar edital"):
            if new_name.strip():
                ed = {
                    "id": new_id(),
                    "name": new_name.strip(),
                    "provaDate": new_date.isoformat() if new_date else "",
                    "weeklyTotalGoal": int(new_goal),
                    "subjects": [],
                }
                editais.append(ed)
                state["currentEditalId"] = ed["id"]
                save_state(state)
                st.rerun()

    if current_edital and st.button("🗑️ Excluir edital atual", use_container_width=True):
        state["editais"] = [e for e in editais if e["id"] != current_edital["id"]]
        state["currentEditalId"] = state["editais"][0]["id"] if state["editais"] else None
        save_state(state)
        st.rerun()

if not current_edital:
    st.info("Crie um edital na barra lateral para começar.")
    st.stop()

# ----------------------------------------------------------------------------
# Barra superior
# ----------------------------------------------------------------------------

top_l, top_r = st.columns([5, 2])
with top_l:
    st.markdown('<div class="topbar-search">', unsafe_allow_html=True)
    search_query = st.text_input(
        "Buscar", placeholder="Buscar matéria ou assunto…", label_visibility="collapsed", key="search_query"
    )
    st.markdown("</div>", unsafe_allow_html=True)
with top_r:
    c1, c2 = st.columns([4, 1])
    with c1:
        st.markdown(
            f'<div style="text-align:right; padding-top:6px;">'
            f'<span style="color:{MUTED};font-size:12px;">Bem-vindo(a)</span><br>'
            f'<b>{user_label}</b></div>',
            unsafe_allow_html=True,
        )
    with c2:
        st.markdown(f'<div class="avatar-chip">{user_label[0].upper()}</div>', unsafe_allow_html=True)

# ----------------------------------------------------------------------------
# Cabeçalho do edital
# ----------------------------------------------------------------------------

st.markdown(f"# {current_edital['name']}")
remaining = days_until(current_edital.get("provaDate"))
if remaining is not None:
    st.caption(f"⏳ {remaining} dias até a prova" if remaining >= 0 else "✅ Prova já realizada")

subjects = current_edital.get("subjects", [])
total_goal = current_edital.get("weeklyTotalGoal", 0)
sum_weights = sum(s.get("peso", 1) for s in subjects) or 1

if search_query:
    q = search_query.lower()
    visible_subjects = [s for s in subjects if q in s["materia"].lower() or q in s.get("assunto", "").lower()]
else:
    visible_subjects = subjects

page = st.session_state.page

# ----------------------------------------------------------------------------
# Página: Painel
# ----------------------------------------------------------------------------

if page == "Painel":
    offset_key = f"week_offset_{current_edital['id']}"
    st.session_state.setdefault(offset_key, 0)
    offset = st.session_state[offset_key]
    wk = week_info(offset)
    week_data = state["weeklyLog"].get(wk["key"], {}).get(current_edital["id"], {})

    if not total_goal:
        st.warning("Defina a meta semanal total na aba **Matérias & Pesos**.")
    elif not subjects:
        st.warning("Cadastre matérias na aba **Matérias & Pesos** ou importe uma planilha.")
    else:
        # ---- coach: recomendação única, só na semana atual ----
        if offset == 0:
            dow = date.today().isoweekday()
            max_weight = max(s.get("peso", 1) for s in subjects)
            best = None
            for s in subjects:
                key = subject_key(s["materia"], s.get("assunto", ""))
                goal = weekly_goal_for(s, total_goal, sum_weights)
                entry = week_data.get(key, {})
                done = entry.get("done", 0)
                correct = entry.get("correct", 0)
                accuracy = correct / done if done > 0 else None
                expected = goal * (dow / 7)
                deficit_ratio = max(0, (expected - done) / goal) if goal else 0
                weight_norm = s.get("peso", 1) / max_weight
                accuracy_factor = 0.5 if accuracy is None else (1 - accuracy)
                score = weight_norm * 0.4 + deficit_ratio * 0.45 + accuracy_factor * 0.15
                reason = "alta prioridade no edital"
                if deficit_ratio > 0.35:
                    reason = "atrasado em relação à meta semanal"
                elif accuracy is not None and accuracy < 0.6:
                    reason = "aproveitamento baixo — vale reforçar"
                if best is None or score > best["score"]:
                    best = {"subject": s, "score": score, "reason": reason}

            if best:
                s = best["subject"]
                st.markdown(
                    f"""<div class="coach-card">
                        <div class="coach-eyebrow">🎯 ESTUDE AGORA</div>
                        <div class="coach-subject">{s['materia']}</div>
                        <div class="coach-topic">{s.get('assunto','')}</div>
                        <span class="coach-reason">{best['reason']}</span>
                    </div>""",
                    unsafe_allow_html=True,
                )
                if s.get("link"):
                    st.link_button("Abrir no TEC ↗", s["link"])

        total_done = sum(week_data.get(subject_key(s["materia"], s.get("assunto", "")), {}).get("done", 0) for s in subjects)
        total_correct = sum(week_data.get(subject_key(s["materia"], s.get("assunto", "")), {}).get("correct", 0) for s in subjects)
        total_goal_sum = sum(weekly_goal_for(s, total_goal, sum_weights) for s in subjects)
        pct = round(100 * total_done / total_goal_sum) if total_goal_sum else 0
        accuracy_pct = round(100 * total_correct / total_done) if total_done else 0
        streak = compute_streak(state["dailyActivity"])

        prev_wk = week_info(-1)
        prev_data = state["weeklyLog"].get(prev_wk["key"], {}).get(current_edital["id"], {})
        prev_done = sum(prev_data.get(subject_key(s["materia"], s.get("assunto", "")), {}).get("done", 0) for s in subjects)
        delta_txt = None
        if prev_done:
            delta_pct = round(100 * (total_done - prev_done) / prev_done)
            delta_txt = f"{'+' if delta_pct >= 0 else ''}{delta_pct}% vs. semana anterior"

        col_main, col_side = st.columns([3, 1])

        with col_main:
            with st.container(border=True):
                nav_l, nav_c, nav_r = st.columns([1, 5, 1])
                with nav_l:
                    if st.button("◀", key="prev_week"):
                        st.session_state[offset_key] -= 1
                        st.rerun()
                with nav_r:
                    if st.button("▶", key="next_week"):
                        st.session_state[offset_key] += 1
                        st.rerun()
                with nav_c:
                    st.markdown(f"**Semana {wk['label']}**" + (" · atual" if offset == 0 else ""))

                st.markdown('<div class="hero-eyebrow">QUESTÕES RESPONDIDAS NA SEMANA</div>', unsafe_allow_html=True)
                delta_html = f'<span class="hero-delta">{delta_txt}</span>' if delta_txt else ""
                st.markdown(
                    f'<div class="hero-number">{total_done} / {total_goal_sum}{delta_html}</div>',
                    unsafe_allow_html=True,
                )
                st.progress(min(1.0, total_done / total_goal_sum) if total_goal_sum else 0.0)

                rows = []
                for o in range(-7, 1):
                    wk_o = week_info(o)
                    ed_log = state["weeklyLog"].get(wk_o["key"], {}).get(current_edital["id"], {})
                    done_o = sum(ed_log.get(subject_key(s["materia"], s.get("assunto", "")), {}).get("done", 0) for s in subjects)
                    goal_o = total_goal_sum
                    rows.append({"Semana": wk_o["monday"].strftime("%d/%m"), "Meta": goal_o, "Feitas": done_o})
                hist_df = pd.DataFrame(rows).set_index("Semana")
                st.bar_chart(hist_df[["Feitas"]], color=ACCENT)

            m1, m2, m3 = st.columns(3)
            with m1:
                with st.container(border=True):
                    st.markdown(f'<span class="metric-dot" style="background:{ACCENT}"></span>**Questões feitas**', unsafe_allow_html=True)
                    st.markdown(f'<div class="hero-number" style="font-size:26px;">{total_done}</div>', unsafe_allow_html=True)
                    st.progress(min(1.0, total_done / total_goal_sum) if total_goal_sum else 0.0)
            with m2:
                with st.container(border=True):
                    st.markdown(f'<span class="metric-dot" style="background:{GREEN}"></span>**Aproveitamento**', unsafe_allow_html=True)
                    st.markdown(f'<div class="hero-number" style="font-size:26px;">{accuracy_pct}%</div>', unsafe_allow_html=True)
                    st.progress(accuracy_pct / 100)
            with m3:
                with st.container(border=True):
                    st.markdown(f'<span class="metric-dot" style="background:{PURPLE}"></span>**Streak**', unsafe_allow_html=True)
                    st.markdown(f'<div class="hero-number" style="font-size:26px;">{streak} dias</div>', unsafe_allow_html=True)
                    st.progress(min(1.0, streak / 30))

            st.markdown("#### Matérias")
            class_filter = st.radio(
                "Filtrar", ["Todas", "Geral", "Específico"], horizontal=True,
                label_visibility="collapsed", key="class_filter",
            )
            if class_filter != "Todas":
                visible_subjects_shown = [s for s in visible_subjects if s.get("classificacao", "Geral") == class_filter]
            else:
                visible_subjects_shown = visible_subjects
            if not visible_subjects_shown:
                st.caption("Nenhuma matéria encontrada para esse filtro.")

            for idx, s in enumerate(visible_subjects_shown):
                key = subject_key(s["materia"], s.get("assunto", ""))
                widget_id = f"{idx}_{key}"
                goal = weekly_goal_for(s, total_goal, sum_weights)
                entry = week_data.get(key, {})
                done = entry.get("done", 0)
                correct = entry.get("correct", 0)
                color = subject_color(s["materia"])
                initials = "".join(w[0] for w in s["materia"].split()[:2]).upper() or "?"
                classificacao = s.get("classificacao", "Geral")
                class_css = "class-especifico" if classificacao == "Específico" else "class-geral"

                with st.container(border=True):
                    icon_col, info_col, input_col = st.columns([1, 4, 3])
                    with icon_col:
                        st.markdown(
                            f'<div class="subject-icon" style="background:{color};">{initials}</div>',
                            unsafe_allow_html=True,
                        )
                    with info_col:
                        st.markdown(
                            f"**{s['materia']}**" + (f" · _{s['assunto']}_" if s.get("assunto") else "")
                            + f' <span class="class-badge {class_css}">{classificacao}</span>',
                            unsafe_allow_html=True,
                        )
                        if goal:
                            st.progress(min(1.0, done / goal), text=f"{done} / {goal} questões")
                    with input_col:
                        colA, colB, colC = st.columns([1, 1, 2])
                        with colA:
                            new_done = st.number_input(
                                "Feitas", min_value=0, value=int(done),
                                key=f"done_{current_edital['id']}_{wk['key']}_{widget_id}",
                            )
                        with colB:
                            new_correct = st.number_input(
                                "Acertos", min_value=0, max_value=max(int(new_done), 0),
                                value=min(int(correct), int(new_done)),
                                key=f"corr_{current_edital['id']}_{wk['key']}_{widget_id}",
                            )
                        with colC:
                            if new_done:
                                st.metric("Aproveitamento", f"{round(100*new_correct/new_done)}%")
                            if s.get("link"):
                                st.link_button("Estudar no TEC ↗", s["link"])

                    if s.get("link"):
                        with st.expander("Tentar abrir aqui dentro do site"):
                            st.components.v1.iframe(s["link"], height=500)
                            st.caption(
                                "Se a tela acima aparecer em branco, o TEC Concursos está bloqueando a "
                                "incorporação por segurança — use o botão 'Estudar no TEC' acima, que abre em nova guia."
                            )

                    with st.expander("📋 Colar contador do TEC"):
                        st.caption(
                            "Copie o textinho do TEC (ex.: **1 de 964 (1 R, 1 A e 0 E)**) e cole abaixo — "
                            "o app extrai Feitas e Acertos automaticamente."
                        )
                        paste_key = f"tec_paste_{current_edital['id']}_{wk['key']}_{widget_id}"
                        pasted = st.text_input("Colar aqui", key=paste_key, placeholder="1 de 964 (1 R, 1 A e 0 E)")
                        if st.button("Aplicar", key=f"apply_{paste_key}"):
                            parsed = parse_tec_counter(pasted)
                            if not parsed:
                                st.error("Não reconheci esse formato. Confira se copiou o trecho com R, A e E.")
                            else:
                                p_done, p_correct, _p_wrong = parsed
                                wk_log = state["weeklyLog"].setdefault(wk["key"], {})
                                ed_log = wk_log.setdefault(current_edital["id"], {})
                                if offset == 0 and p_done > done:
                                    today_str = date.today().isoformat()
                                    state["dailyActivity"][today_str] = state["dailyActivity"].get(today_str, 0) + (p_done - done)
                                ed_log[key] = {"done": p_done, "correct": min(p_correct, p_done)}
                                save_state(state)
                                st.rerun()

                    if new_done != done or new_correct != correct:
                        wk_log = state["weeklyLog"].setdefault(wk["key"], {})
                        ed_log = wk_log.setdefault(current_edital["id"], {})
                        ed_log[key] = {"done": int(new_done), "correct": int(min(new_correct, new_done))}
                        if offset == 0 and new_done > done:
                            today_str = date.today().isoformat()
                            state["dailyActivity"][today_str] = state["dailyActivity"].get(today_str, 0) + (new_done - done)
                        save_state(state)

        with col_side:
            st.markdown(
                f"""<div class="today-card">
                    <div class="hero-eyebrow">PROGRESSO DA META SEMANAL</div>
                    <div class="hero-number">{pct}%</div>
                    <div style="color:{MUTED}; font-size:13px; margin-top:4px;">
                        {total_done} de {total_goal_sum} questões nesta semana
                    </div>
                </div>""",
                unsafe_allow_html=True,
            )
            if st.button("Ver dashboard completo →", use_container_width=True):
                st.session_state.page = "Dashboard"
                st.rerun()

            with st.container(border=True):
                st.markdown("**Aproveitamento por matéria**")
                perf_rows = []
                for s in subjects:
                    k = subject_key(s["materia"], s.get("assunto", ""))
                    v = week_data.get(k, {})
                    d = v.get("done", 0)
                    c_ = v.get("correct", 0)
                    if d:
                        perf_rows.append({"Matéria": s["materia"], "Acerto %": round(100 * c_ / d)})
                if perf_rows:
                    perf_df = pd.DataFrame(perf_rows).groupby("Matéria").mean().sort_values("Acerto %", ascending=False).head(6)
                    st.bar_chart(perf_df, color=GREEN)
                else:
                    st.caption("Registre questões feitas para ver o aproveitamento por matéria.")

# ----------------------------------------------------------------------------
# Página: Matérias & Pesos
# ----------------------------------------------------------------------------

elif page == "Matérias & Pesos":
    new_goal_val = st.number_input(
        "Meta semanal total de questões", min_value=0, value=int(total_goal), key=f"goal_{current_edital['id']}"
    )
    if int(new_goal_val) != total_goal:
        current_edital["weeklyTotalGoal"] = int(new_goal_val)
        save_state(state)
        st.rerun()

    st.caption(
        "Edite direto na tabela. Para adicionar uma matéria, use a última linha em branco. "
        "Para excluir, selecione a linha e aperte a lixeira que aparece no canto."
    )

    df = pd.DataFrame(subjects) if subjects else pd.DataFrame(columns=["materia", "assunto", "peso", "classificacao", "link"])
    if "classificacao" not in df.columns:
        df["classificacao"] = "Geral"
    df["classificacao"] = df["classificacao"].fillna("Geral")
    df = df[["materia", "assunto", "peso", "classificacao", "link"]]
    df = df.rename(columns={
        "materia": "Matéria", "assunto": "Assunto", "peso": "Peso",
        "classificacao": "Classificação", "link": "Link TEC",
    })

    edited = st.data_editor(
        df,
        num_rows="dynamic",
        use_container_width=True,
        key=f"editor_{current_edital['id']}",
        column_config={
            "Peso": st.column_config.NumberColumn("Peso (1 a 5)", min_value=1, max_value=5, step=0.5),
            "Classificação": st.column_config.SelectboxColumn("Classificação", options=["Geral", "Específico"], default="Geral"),
            "Link TEC": st.column_config.LinkColumn("Link TEC"),
        },
    )

    new_subjects = []
    for _, row in edited.iterrows():
        materia = str(row.get("Matéria", "")).strip()
        if not materia or materia.lower() == "nan":
            continue
        new_subjects.append(
            {
                "materia": materia,
                "assunto": "" if pd.isna(row.get("Assunto")) else str(row.get("Assunto")).strip(),
                "peso": float(row.get("Peso")) if pd.notna(row.get("Peso")) else 1.0,
                "classificacao": row.get("Classificação") if row.get("Classificação") in ("Geral", "Específico") else "Geral",
                "link": "" if pd.isna(row.get("Link TEC")) else str(row.get("Link TEC")).strip(),
            }
        )

    if new_subjects != subjects:
        current_edital["subjects"] = new_subjects
        save_state(state)

    if new_subjects:
        sw = sum(s["peso"] for s in new_subjects) or 1
        preview = pd.DataFrame(
            [
                {
                    "Matéria": s["materia"],
                    "Peso": s["peso"],
                    "Meta semanal calculada": weekly_goal_for(s, current_edital.get("weeklyTotalGoal", 0), sw),
                }
                for s in new_subjects
            ]
        )
        st.markdown("**Distribuição calculada da meta semanal:**")
        st.dataframe(preview, use_container_width=True, hide_index=True)

# ----------------------------------------------------------------------------
# Página: Importar planilha
# ----------------------------------------------------------------------------

elif page == "Importar planilha":
    st.caption(
        "Padrão de upload: sua planilha de **Edital Verticalizado** (colunas Grupo/Disciplina, "
        "Conteúdo Programático/Tópico, Classificação Geral/Específico e, opcionalmente, Link TEC). "
        "Também aceita uma planilha simples com colunas Materia/Assunto/Peso/Classificação/LinkTEC."
    )

    st.download_button(
        "⬇️ Baixar planilha modelo (.xlsx)",
        build_template_xlsx(),
        "modelo-edital-verticalizado.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    file = st.file_uploader("Envie a planilha do edital", type=["csv", "xlsx", "xls"])

    if file:
        try:
            imported = None
            is_csv = file.name.lower().endswith(".csv")

            if not is_csv:
                xls = pd.ExcelFile(file)
                sheet_name = xls.sheet_names[0]
                if len(xls.sheet_names) > 1:
                    sheet_name = st.selectbox("Escolha a aba/cargo desta planilha", xls.sheet_names)
                df_raw = pd.read_excel(xls, sheet_name=sheet_name, header=None)
                imported = parse_verticalizado(df_raw)

            if imported is None:
                # fallback: planilha simples, com cabeçalho na primeira linha
                file.seek(0)
                df_in = pd.read_csv(file) if is_csv else pd.read_excel(file, sheet_name=sheet_name if not is_csv else None)
                cols = {_norm(c): c for c in df_in.columns}

                def col(*names):
                    for n in names:
                        if n in cols:
                            return cols[n]
                    return None

                c_mat = col("materia", "disciplina", "grupo")
                c_ass = col("assunto", "topico", "tema", "conteudo programatico")
                c_peso = col("peso", "importancia", "prioridade")
                c_class = col("classificacao", "tipo", "categoria", "geral/especifico")
                c_link = col("linktec", "link tec", "link", "tec")

                if not c_mat:
                    imported = []
                else:
                    imported = []
                    for _, row in df_in.iterrows():
                        materia = str(row[c_mat]).strip()
                        if not materia or materia.lower() == "nan":
                            continue
                        imported.append(
                            {
                                "materia": materia,
                                "assunto": str(row[c_ass]).strip() if c_ass and pd.notna(row[c_ass]) else "",
                                "peso": float(row[c_peso]) if c_peso and pd.notna(row[c_peso]) else 3.0,
                                "classificacao": _normalize_classificacao(row[c_class]) if c_class and pd.notna(row[c_class]) else "Geral",
                                "link": str(row[c_link]).strip() if c_link and pd.notna(row[c_link]) else "",
                                "done": 0,
                                "correct": 0,
                            }
                        )

            if not imported:
                st.error(
                    "Não reconheci o formato da planilha. Confira se ela segue o padrão do "
                    "Edital Verticalizado (Grupo/Disciplina + Conteúdo Programático) ou o formato "
                    "simples com colunas Materia/Assunto."
                )
            else:
                st.markdown(f"**{len(imported)} matérias/assuntos encontrados:**")
                st.dataframe(
                    pd.DataFrame(imported)[["materia", "assunto", "peso", "classificacao", "link"]].rename(
                        columns={
                            "materia": "Matéria", "assunto": "Assunto", "peso": "Peso",
                            "classificacao": "Classificação", "link": "Link TEC",
                        }
                    ),
                    use_container_width=True,
                    hide_index=True,
                )

                has_progress = any(r.get("done") for r in imported)
                seed_progress = False
                if has_progress:
                    seed_progress = st.checkbox(
                        "Também importar Questões Feitas / % Acertos como progresso da semana atual",
                        value=True,
                    )

                def _apply(new_list):
                    current_edital["subjects"] = new_list
                    if seed_progress:
                        wk_now = week_info(0)
                        wk_log = state["weeklyLog"].setdefault(wk_now["key"], {})
                        ed_log = wk_log.setdefault(current_edital["id"], {})
                        for r in imported:
                            if r.get("done"):
                                k = subject_key(r["materia"], r.get("assunto", ""))
                                ed_log[k] = {"done": r["done"], "correct": r.get("correct", 0)}
                    save_state(state)
                    st.rerun()

                clean_imported = [
                    {
                        "materia": r["materia"], "assunto": r["assunto"], "peso": r["peso"],
                        "classificacao": r.get("classificacao", "Geral"), "link": r["link"],
                    }
                    for r in imported
                ]

                colA, colB = st.columns(2)
                with colA:
                    if st.button("➕ Adicionar às matérias existentes"):
                        _apply(subjects + clean_imported)
                with colB:
                    if st.button("♻️ Substituir matérias deste edital"):
                        _apply(clean_imported)
        except Exception as e:
            st.error(f"Erro ao ler o arquivo: {e}")

# ----------------------------------------------------------------------------
# Página: Dashboard
# ----------------------------------------------------------------------------

elif page == "Dashboard":
    total_attempts = 0
    total_correct = 0
    for wk_log in state["weeklyLog"].values():
        ed_log = wk_log.get(current_edital["id"], {})
        for v in ed_log.values():
            total_attempts += v.get("done", 0)
            total_correct += v.get("correct", 0)

    xp = total_attempts * 2 + total_correct * 3
    level, progress = level_from_xp(xp)
    streak = compute_streak(state["dailyActivity"])
    accuracy = total_correct / total_attempts if total_attempts else 0

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Nível", level)
    c2.metric("XP total", xp)
    c3.metric("Streak", f"{streak} dias")
    c4.metric("Aproveitamento geral", f"{round(accuracy*100)}%")
    st.progress(progress)

    badge_defs = [
        ("10 questões", total_attempts >= 10),
        ("100 questões", total_attempts >= 100),
        ("500 questões", total_attempts >= 500),
        ("3 dias seguidos", streak >= 3),
        ("7 dias seguidos", streak >= 7),
        ("30 dias seguidos", streak >= 30),
        ("80% de acerto", total_attempts >= 30 and accuracy >= 0.8),
    ]
    badges_html = "".join(
        f'<span class="{"badge-earned" if earned else "badge-locked"}">🏅 {label}</span>' for label, earned in badge_defs
    )
    st.markdown(badges_html, unsafe_allow_html=True)

    st.markdown("### Evolução — últimas 8 semanas")
    rows = []
    for o in range(-7, 1):
        wk_o = week_info(o)
        ed_log = state["weeklyLog"].get(wk_o["key"], {}).get(current_edital["id"], {})
        done = 0
        correct = 0
        goal = 0
        for s in subjects:
            key = subject_key(s["materia"], s.get("assunto", ""))
            goal += weekly_goal_for(s, total_goal, sum_weights)
            v = ed_log.get(key, {})
            done += v.get("done", 0)
            correct += v.get("correct", 0)
        rows.append(
            {
                "Semana": wk_o["monday"].strftime("%d/%m"),
                "Meta": goal,
                "Feitas": done,
                "Acerto %": round(100 * correct / done) if done else 0,
            }
        )
    hist_df = pd.DataFrame(rows).set_index("Semana")
    st.line_chart(hist_df)

    st.markdown("### Desempenho por matéria (semana atual)")
    week_now = state["weeklyLog"].get(week_info(0)["key"], {}).get(current_edital["id"], {})
    perf_rows = []
    for s in subjects:
        key = subject_key(s["materia"], s.get("assunto", ""))
        v = week_now.get(key, {})
        d = v.get("done", 0)
        c_ = v.get("correct", 0)
        perf_rows.append(
            {
                "Matéria": s["materia"],
                "Peso": s["peso"],
                "Feitas/Meta": f"{d}/{weekly_goal_for(s, total_goal, sum_weights)}",
                "Acerto": f"{round(100*c_/d)}%" if d else "—",
            }
        )
    st.dataframe(pd.DataFrame(perf_rows), use_container_width=True, hide_index=True)
